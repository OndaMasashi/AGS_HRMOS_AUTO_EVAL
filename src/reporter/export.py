"""レポート出力モジュール - AI評価結果をExcel形式で出力"""

import json
import logging
from collections import OrderedDict
from datetime import datetime
from pathlib import Path

from src.evaluator.prompt_builder import total_to_rank, is_first_pass_candidate

logger = logging.getLogger(__name__)


def export_evaluation_excel(
    evaluations: list[dict],
    criteria_names: list[str],
    report_dir: str,
    question_count: int = 3,
    first_pass_criteria: list[dict] | None = None,
) -> str:
    """AI評価結果をExcelファイルに出力する

    1行 = 1応募者
    レイアウト:
      基本情報 → 1次通過候補 → 平均点 → 合計点 → 総合ランク → レーダーチャート → 総合評価
      → 各評価基準(点・評価) → 備考欄 → 質問候補1〜N
    """
    from openpyxl import Workbook
    from openpyxl.chart import RadarChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    Path(report_dir).mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"ai_evaluation_{timestamp}.xlsx"
    filepath = Path(report_dir) / filename

    wb = Workbook()
    ws = wb.active
    ws.title = "AI評価結果"

    # ヘッダー構築（新レイアウト）
    headers = ["応募者名", "性別", "年齢", "HRMOS URL", "ファイル名"]
    headers.append("1次通過候補")
    headers.append("平均点")
    headers.append("合計点")
    headers.append("総合ランク")
    headers.append("レーダーチャート")
    headers.append("総合評価")
    for criteria in criteria_names:
        headers.append(f"{criteria}(点)")
        headers.append(f"{criteria}(評価)")
    headers.append("備考欄")
    for i in range(1, question_count + 1):
        headers.append(f"質問候補{i}")

    # ランク別の色定義
    rank_colors = {
        "S": PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),  # 金
        "A": PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid"),  # 水色
        "B": PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),  # 薄緑
        "C": PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid"),  # 薄橙
        "D": PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"),  # 薄赤
    }

    # スタイル定義
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    score_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    question_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    candidate_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    near_candidate_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")  # 薄黄
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # ヘッダー書き込み
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    # 評価結果を応募者ごとにグルーピング
    by_applicant = OrderedDict()
    for ev in evaluations:
        app_id = ev["applicant_id"]
        if app_id not in by_applicant:
            questions = []
            if ev.get("interview_questions"):
                try:
                    questions = json.loads(ev["interview_questions"])
                except (json.JSONDecodeError, TypeError):
                    questions = []

            by_applicant[app_id] = {
                "name": ev["applicant_name"],
                "gender": ev.get("applicant_gender", "不明"),
                "age": ev.get("applicant_age"),
                "page_url": ev["page_url"],
                "filename": ev["filename"],
                "total_score": ev["total_score"],
                "overall_comment": ev["overall_comment"],
                "remarks": ev.get("remarks", ""),
                "interview_questions": questions,
                "criteria": {}
            }
        by_applicant[app_id]["criteria"][ev["criteria_name"]] = {
            "score": ev["score"],
            "comment": ev["comment"],
        }

    # レーダーチャート用ヘルパーシート作成
    ws_radar = wb.create_sheet("_radar_data")
    for i, name in enumerate(criteria_names, 2):
        ws_radar.cell(row=1, column=i, value=name)

    # 列インデックス定数（1始まり）
    COL_NAME = 1
    COL_GENDER = 2
    COL_AGE = 3
    COL_URL = 4
    COL_FILENAME = 5
    COL_CANDIDATE = 6
    COL_AVG = 7
    COL_TOTAL = 8
    COL_RANK = 9
    COL_RADAR = 10
    COL_OVERALL = 11
    COL_CRITERIA_START = 12
    # 備考欄の列 = COL_CRITERIA_START + len(criteria_names) * 2
    # 質問候補の列 = 備考欄 + 1

    # データ行書き込み
    for row_idx, (app_id, app_data) in enumerate(by_applicant.items(), 2):
        # 応募者名
        ws.cell(row=row_idx, column=COL_NAME, value=app_data["name"]).border = thin_border

        # 性別
        gender_cell = ws.cell(row=row_idx, column=COL_GENDER, value=app_data.get("gender", "不明"))
        gender_cell.alignment = Alignment(horizontal="center")
        gender_cell.border = thin_border

        # 年齢
        age_val = app_data.get("age")
        age_cell = ws.cell(row=row_idx, column=COL_AGE, value=age_val if age_val is not None else "不明")
        age_cell.alignment = Alignment(horizontal="center")
        age_cell.border = thin_border

        # HRMOS URL
        ws.cell(row=row_idx, column=COL_URL, value=app_data["page_url"]).border = thin_border

        # ファイル名
        filename_cell = ws.cell(row=row_idx, column=COL_FILENAME, value=app_data["filename"])
        filename_cell.alignment = Alignment(wrap_text=True, vertical="top")
        filename_cell.border = thin_border

        # 平均点・1次通過候補の計算
        criteria_count = len(criteria_names)
        avg_score = round(app_data["total_score"] / criteria_count, 1) if criteria_count > 0 else 0
        candidate_mark = is_first_pass_candidate(avg_score, age_val, first_pass_criteria or [])

        # 1次通過候補
        candidate_cell = ws.cell(row=row_idx, column=COL_CANDIDATE, value=candidate_mark)
        candidate_cell.alignment = Alignment(horizontal="center", vertical="center")
        candidate_cell.border = thin_border
        candidate_cell.font = Font(bold=True, size=14)
        if candidate_mark == "○":
            candidate_cell.fill = candidate_fill
        elif candidate_mark == "△":
            candidate_cell.fill = near_candidate_fill

        # 平均点
        avg_cell = ws.cell(row=row_idx, column=COL_AVG, value=avg_score)
        avg_cell.font = Font(bold=True)
        avg_cell.alignment = Alignment(horizontal="center")
        avg_cell.border = thin_border

        # 合計点
        total_cell = ws.cell(row=row_idx, column=COL_TOTAL, value=app_data["total_score"])
        total_cell.font = Font(bold=True)
        total_cell.alignment = Alignment(horizontal="center")
        total_cell.border = thin_border

        # 総合ランク
        rank = total_to_rank(app_data["total_score"], criteria_count)
        rank_cell = ws.cell(row=row_idx, column=COL_RANK, value=rank)
        rank_cell.font = Font(bold=True, size=12)
        rank_cell.alignment = Alignment(horizontal="center", vertical="center")
        rank_cell.border = thin_border
        if rank in rank_colors:
            rank_cell.fill = rank_colors[rank]

        # レーダーチャート列（空セル、チャートは後で追加）
        radar_cell = ws.cell(row=row_idx, column=COL_RADAR, value="")
        radar_cell.border = thin_border

        # 総合評価
        overall_cell = ws.cell(row=row_idx, column=COL_OVERALL, value=app_data["overall_comment"])
        overall_cell.alignment = Alignment(wrap_text=True, vertical="top")
        overall_cell.border = thin_border

        # 各評価基準
        col = COL_CRITERIA_START
        for criteria in criteria_names:
            c_data = app_data["criteria"].get(criteria, {"score": 0, "comment": "未評価"})

            score_cell = ws.cell(row=row_idx, column=col, value=c_data["score"])
            score_cell.fill = score_fill
            score_cell.alignment = Alignment(horizontal="center")
            score_cell.border = thin_border
            col += 1

            comment_cell = ws.cell(row=row_idx, column=col, value=c_data["comment"])
            comment_cell.alignment = Alignment(wrap_text=True, vertical="top")
            comment_cell.border = thin_border
            col += 1

        # 備考欄
        remarks_cell = ws.cell(row=row_idx, column=col, value=app_data.get("remarks", ""))
        remarks_cell.alignment = Alignment(wrap_text=True, vertical="top")
        remarks_cell.border = thin_border
        col += 1

        # 質問候補
        questions = app_data.get("interview_questions", [])
        for i in range(question_count):
            q_text = questions[i] if i < len(questions) else ""
            q_cell = ws.cell(row=row_idx, column=col, value=q_text)
            q_cell.fill = question_fill
            q_cell.alignment = Alignment(wrap_text=True, vertical="top")
            q_cell.border = thin_border
            col += 1

        # レーダーチャート用データをヘルパーシートに書き込み
        ws_radar.cell(row=row_idx, column=1, value=app_data["name"])
        for i, criteria in enumerate(criteria_names, 2):
            score = app_data["criteria"].get(criteria, {}).get("score", 0)
            ws_radar.cell(row=row_idx, column=i, value=score)

        # 行の高さを拡大（レーダーチャート用）
        ws.row_dimensions[row_idx].height = 120

    # レーダーチャート生成
    for r_idx in range(2, 2 + len(by_applicant)):
        chart = RadarChart()
        chart.type = "filled"
        chart.style = 10
        chart.width = 3.5    # cm
        chart.height = 3.5   # cm
        chart.title = None
        chart.legend = None

        data_ref = Reference(
            ws_radar,
            min_col=2, max_col=1 + len(criteria_names),
            min_row=r_idx, max_row=r_idx
        )
        cats_ref = Reference(
            ws_radar,
            min_col=2, max_col=1 + len(criteria_names),
            min_row=1
        )
        chart.add_data(data_ref, from_rows=True)
        chart.set_categories(cats_ref)

        # シリーズのスタイル設定（フラットな塗りつぶし）
        if chart.series:
            from openpyxl.drawing.fill import GradientFillProperties
            s = chart.series[0]
            s.graphicalProperties.line.width = 20000  # 1.5pt
            s.graphicalProperties.line.solidFill = "2F5496"  # 濃い青の線
            s.graphicalProperties.solidFill = "4472C4"       # 青系の塗りつぶし

        # レーダーチャートの軸スケールを固定（0-5）
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 5

        anchor_cell = f"{get_column_letter(COL_RADAR)}{r_idx}"
        ws.add_chart(chart, anchor_cell)

    # ヘルパーシートを非表示
    ws_radar.sheet_state = 'hidden'

    # 列幅設定
    ws.column_dimensions[get_column_letter(COL_NAME)].width = 15       # 応募者名
    ws.column_dimensions[get_column_letter(COL_GENDER)].width = 6      # 性別
    ws.column_dimensions[get_column_letter(COL_AGE)].width = 6         # 年齢
    ws.column_dimensions[get_column_letter(COL_URL)].width = 15        # HRMOS URL（縮小）
    ws.column_dimensions[get_column_letter(COL_FILENAME)].width = 12   # ファイル名（縮小）
    ws.column_dimensions[get_column_letter(COL_CANDIDATE)].width = 10  # 1次通過候補
    ws.column_dimensions[get_column_letter(COL_AVG)].width = 8         # 平均点
    ws.column_dimensions[get_column_letter(COL_TOTAL)].width = 10      # 合計点
    ws.column_dimensions[get_column_letter(COL_RANK)].width = 10       # 総合ランク
    ws.column_dimensions[get_column_letter(COL_RADAR)].width = 18      # レーダーチャート
    ws.column_dimensions[get_column_letter(COL_OVERALL)].width = 50    # 総合評価

    for i in range(len(criteria_names)):
        score_col = get_column_letter(COL_CRITERIA_START + i * 2)
        comment_col = get_column_letter(COL_CRITERIA_START + i * 2 + 1)
        ws.column_dimensions[score_col].width = 8
        ws.column_dimensions[comment_col].width = 40

    remarks_col_idx = COL_CRITERIA_START + len(criteria_names) * 2
    ws.column_dimensions[get_column_letter(remarks_col_idx)].width = 40  # 備考欄

    for i in range(question_count):
        ws.column_dimensions[get_column_letter(remarks_col_idx + 1 + i)].width = 45

    ws.freeze_panes = "A2"

    wb.save(filepath)
    logger.info(f"AI評価Excelレポート出力: {filepath} ({len(by_applicant)} 名)")
    return str(filepath)
